"""
Valuation engine for the private credit quarter-end app (DUMMY DATA).

Every loader accepts either a filesystem path OR an uploaded file / bytes buffer,
so the Streamlit app can run entirely off user-uploaded workbooks:

  load_market_data(src)      - Market_Data_Inputs.xlsx  -> {'Q1': {...}, 'Q2': {...}}
  load_model(src)            - PC_Valuation_Qx_2026.xlsx -> positions, recovery params,
                               prior-quarter marks (Bridge tab), valuation date

Math mirrors the Excel models:
  market yield = spot SOFR + credit spread(rating x seniority) + illiquidity premium + credit adj
  quarterly contractual CFs off the SOFR forward curve (with floors); PIK capitalizes; bullet maturity.
"""
from __future__ import annotations
import datetime as dt
import io
import math
from dataclasses import dataclass

import pandas as pd
from openpyxl import load_workbook

QUARTER_END = {"Q1": dt.date(2026, 3, 31), "Q2": dt.date(2026, 6, 30)}


def _wb(src):
    """Accept a path, bytes, or a file-like object (e.g. Streamlit UploadedFile)."""
    if isinstance(src, (bytes, bytearray)):
        src = io.BytesIO(src)
    elif hasattr(src, "seek"):
        src.seek(0)
    return load_workbook(src, data_only=True)


def _d(v):
    return v.date() if isinstance(v, dt.datetime) else v


# ----------------------------------------------------------------- market data
def load_market_data(src) -> dict:
    """Parse Market_Data_Inputs.xlsx into a dict keyed by vintage ('Q1'/'Q2')."""
    wb = _wb(src)
    if "Market_Data" not in wb.sheetnames:
        raise ValueError("This doesn't look like Market_Data_Inputs.xlsx (no 'Market_Data' tab).")
    ws = wb["Market_Data"]
    out = {}
    for j, v in enumerate(("Q1", "Q2")):
        col = "BC"[j]
        spot = ws[f"{col}3"].value
        fwd = [ws[f"{col}{8+i}"].value for i in range(20)]
        base = 30 + j * 8
        matrix = {ws[f"A{base+2+si}"].value: [ws.cell(row=base + 2 + si, column=2 + k).value for k in range(5)]
                  for si in range(4)}
        illiq = {ws[f"A{47+si}"].value: ws[f"B{47+si}"].value for si in range(4)}
        bench = {ws[f"A{55+bi}"].value: ws[f"{col}{55+bi}"].value for bi in range(4)}
        if spot is None or any(x is None for x in fwd):
            raise ValueError("Market data tab layout not recognized — expected the template "
                             "produced with this app (spot in row 3, forward curve rows 8-27).")
        out[v] = {"spot": spot, "fwd": fwd, "spreads": matrix, "illiq": illiq, "bench": bench}
    return out


# ------------------------------------------------------------- portfolio model
@dataclass
class Position:
    name: str
    industry: str
    seniority: str
    rating: int
    origination: dt.date
    maturity: dt.date
    commitment: float
    funded: float
    cash_spread_bps: float
    pik_bps: float
    floor: float
    oid: float
    credit_adj_bps: float
    recovery_weight: float


def load_model(src) -> dict:
    """Parse a PC_Valuation_Qx_2026.xlsx workbook: positions, recovery params,
    prior-quarter marks (Bridge tab) and the valuation date on the Summary tab."""
    wb = _wb(src)
    for tab in ("Positions", "Recovery", "Bridge", "Summary"):
        if tab not in wb.sheetnames:
            raise ValueError(f"This doesn't look like a valuation model workbook (missing '{tab}' tab).")

    ws = wb["Positions"]
    positions, r = [], 5
    while ws[f"B{r}"].value and ws[f"B{r}"].value != "TOTAL / WTD AVG":
        positions.append(Position(
            name=ws[f"B{r}"].value, industry=ws[f"C{r}"].value, seniority=ws[f"D{r}"].value,
            rating=int(ws[f"E{r}"].value), origination=_d(ws[f"F{r}"].value), maturity=_d(ws[f"G{r}"].value),
            commitment=float(ws[f"H{r}"].value), funded=float(ws[f"I{r}"].value),
            cash_spread_bps=float(ws[f"J{r}"].value), pik_bps=float(ws[f"K{r}"].value),
            floor=float(ws[f"L{r}"].value), oid=float(ws[f"M{r}"].value),
            credit_adj_bps=float(ws[f"R{r}"].value), recovery_weight=float(ws[f"U{r}"].value),
        ))
        r += 1
    if not positions:
        raise ValueError("No positions found on the Positions tab (expected data from row 5).")

    rc = wb["Recovery"]
    recovery = {"ebitda": rc["B4"].value, "multiple": rc["B5"].value, "revolver": rc["B7"].value,
                "fl_claims": rc["B9"].value, "distressed_rate": rc["B13"].value,
                "resolution_yrs": rc["B12"].value}

    br = wb["Bridge"]
    marks, r = [], 4
    while br[f"A{r}"].value and br[f"A{r}"].value != "TOTAL":
        marks.append({"name": br[f"A{r}"].value, "prior_fv": br[f"B{r}"].value or 0.0,
                      "prior_funded": br[f"C{r}"].value or 0.0, "par_repaid": br[f"D{r}"].value or 0.0,
                      "credit_event": br[f"K{r}"].value == "Y"})
        r += 1
    # Bridge column A holds formulas (=Positions!B..); if the file was saved without cached
    # values those read back as None — fall back to position order, which the Bridge tab
    # shares by construction.
    for i, m in enumerate(marks):
        if m["name"] is None and i < len(positions):
            m["name"] = positions[i].name

    valdate = _d(wb["Summary"]["B3"].value)
    return {"positions": positions, "recovery": recovery,
            "prior_marks": pd.DataFrame(marks).set_index("name") if marks else pd.DataFrame(),
            "valdate": valdate}


# ---------------------------------------------------------------------- engine
def n_periods(valdate: dt.date, maturity: dt.date) -> int:
    return max(1, math.ceil((maturity - valdate).days / 365 * 4))


def market_yield(p: Position, md: dict, spread_shock_bps: float = 0.0,
                 include_illiquidity: bool = True) -> dict:
    spread = md["spreads"][p.seniority][p.rating - 1]
    illiq = md["illiq"][p.seniority] if include_illiquidity else 0.0
    total = md["spot"] + (spread + illiq + p.credit_adj_bps + spread_shock_bps) / 10000
    return {"base": md["spot"], "credit_spread_bps": spread, "illiq_bps": illiq,
            "credit_adj_bps": p.credit_adj_bps, "shock_bps": spread_shock_bps, "yield": total}


def project_cashflows(p: Position, valdate: dt.date, md: dict, y: float,
                      pik_capitalize: bool = True) -> pd.DataFrame:
    """Quarterly contractual CFs. pik_capitalize=False treats the PIK coupon as cash-pay."""
    N = n_periods(valdate, p.maturity)
    bal = p.funded
    recs = []
    for n in range(1, N + 1):
        fwd = md["fwd"][min(n, 20) - 1]
        base = max(fwd, p.floor)
        cash_rate = base + p.cash_spread_bps / 10000
        pik_rate = p.pik_bps / 10000
        if pik_capitalize:
            pik_acc = bal * pik_rate / 4
            cash_int = bal * cash_rate / 4
        else:
            pik_acc = 0.0
            cash_int = bal * (cash_rate + pik_rate) / 4
        end_bal = bal + pik_acc
        principal = end_bal if n == N else 0.0
        cf = cash_int + principal
        df = 1 / (1 + y / 4) ** n
        recs.append({"n": n, "period_end": valdate + dt.timedelta(days=round(91.25 * n)),
                     "fwd_sofr": fwd, "base": base, "beg_balance": bal, "pik_accrual": pik_acc,
                     "cash_interest": cash_int, "principal": principal, "total_cf": cf,
                     "df": df, "pv": cf * df})
        bal = end_bal
    return pd.DataFrame(recs)


def recovery_value_pct(rp: dict) -> float:
    ev = rp["ebitda"] * rp["multiple"]
    to_fl = max(0.0, ev - rp["revolver"])
    gross = min(1.0, to_fl / rp["fl_claims"])
    return gross / (1 + rp["distressed_rate"]) ** rp["resolution_yrs"]


def value_portfolio(model: dict, md: dict, valdate: dt.date, pik_capitalize: bool = True,
                    spread_shock_bps: float = 0.0, include_illiquidity: bool = True) -> pd.DataFrame:
    rec_pct = recovery_value_pct(model["recovery"])
    rows = []
    for p in model["positions"]:
        yb = market_yield(p, md, spread_shock_bps, include_illiquidity)
        cfs = project_cashflows(p, valdate, md, yb["yield"], pik_capitalize)
        dcf_fv = cfs["pv"].sum()
        rec_fv = rec_pct * p.funded if p.recovery_weight > 0 else 0.0
        final = (1 - p.recovery_weight) * dcf_fv + p.recovery_weight * rec_fv
        rows.append({"Borrower": p.name, "Industry": p.industry, "Seniority": p.seniority,
                     "Rating": p.rating, "Maturity": p.maturity, "Commitment": p.commitment,
                     "Funded": p.funded, "Cash Spread (bps)": p.cash_spread_bps,
                     "PIK (bps)": p.pik_bps, "Floor": p.floor,
                     "Base": yb["base"], "Credit Spread (bps)": yb["credit_spread_bps"],
                     "Illiquidity (bps)": yb["illiq_bps"], "Credit Adj (bps)": yb["credit_adj_bps"],
                     "Market Yield": yb["yield"], "DCF FV": dcf_fv,
                     "Recovery Weight": p.recovery_weight, "Recovery FV": rec_fv,
                     "Fair Value": final, "FV % of Funded": final / p.funded})
    return pd.DataFrame(rows)


def qoq_bridge(prior_df: pd.DataFrame | None, current_df: pd.DataFrame,
               marks: pd.DataFrame) -> pd.DataFrame:
    """Position-level bridge mirroring the Excel Bridge tab.
    Drivers: repayments, PIK capitalized, new originations, credit-specific, market re-mark."""
    prior_fv_engine = prior_df.set_index("Borrower")["Fair Value"] if prior_df is not None else None
    rows = []
    for _, r in current_df.iterrows():
        nm = r["Borrower"]
        m = marks.loc[nm] if nm in marks.index else None
        key = nm.replace(" (NEW)", "")
        if prior_fv_engine is not None and key in prior_fv_engine.index and "(NEW)" not in nm:
            prior_fv = float(prior_fv_engine[key])  # consistent with the user's toggles
        else:
            prior_fv = float(m["prior_fv"]) if m is not None else 0.0
        prior_funded = float(m["prior_funded"]) if m is not None else 0.0
        repaid = float(m["par_repaid"]) if m is not None else 0.0
        credit_event = bool(m["credit_event"]) if m is not None else False
        cur_fv, funded, pik = r["Fair Value"], r["Funded"], r["PIK (bps)"]
        is_new = prior_fv == 0 and prior_funded == 0
        repay_impact = -repaid * (prior_fv / prior_funded) if prior_funded else 0.0
        pik_cap = (funded - (prior_funded or funded)) * (cur_fv / funded) if pik > 0 else 0.0
        new_orig = cur_fv if is_new else 0.0
        resid = cur_fv - prior_fv - repay_impact - pik_cap - new_orig
        rows.append({"Borrower": nm, "Prior FV": prior_fv, "Repayments": repay_impact,
                     "PIK Capitalized": pik_cap, "New Originations": new_orig,
                     "Credit-Specific": resid if credit_event else 0.0,
                     "Market Re-mark & Accretion": 0.0 if credit_event else resid,
                     "Current FV": cur_fv})
    return pd.DataFrame(rows)
